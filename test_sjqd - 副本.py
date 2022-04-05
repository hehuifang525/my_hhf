"""
    数据驱动的封装和打包
"""

import openpyxl
from openpyxl.styles import PatternFill,Font
from base.logger import Logger


# 读取excel文件
from base.keys import Keys


def sjqd(file_path):
    wb = openpyxl.load_workbook(file_path)

    # 循环读取excel中的内容

    for each_sheet in wb.sheetnames:  # 遍历每一个sheet
        for i in wb[each_sheet].values: # 将每一个sheet的值取出
            # 处理表头第一行
            if type(i[0]) is int:
                # 处理数据
                data1 = {"action": i[1], "parameter": eval(str(i[2])), "describe": i[3]}
                """
                    开始处理操作行为
                    1、打开浏览器，初始化对象
                    2、常规操作行为
                    3、断言-需要比较，以及回写校验结果
                """

                if data1["action"] == "open_browser":
                    keys = Keys(**data1["parameter"])  # 字典解包传入参数
                elif data1["action"] == "quit":
                    getattr(keys, data1["action"])()

                elif "assert" in data1["action"]:
                    try:
                        getattr(keys, data1["action"])(**data1["parameter"])
                        wb[each_sheet].cell(row=i[0]+1, column=5).value="pass"
                        pass_fill = PatternFill(start_color='4FB164', end_color='4FB164', fill_type='solid')
                        pass_font = Font(bold=True)
                        wb[each_sheet].cell(row=i[0] + 1, column=5).fill = pass_fill
                        wb[each_sheet].cell(row=i[0] + 1, column=5).font = pass_font

                    except Exception as e:
                        wb[each_sheet].cell(row=i[0]+1, column=5).value = "fail"
                        fail_fill = PatternFill(start_color='F3490D', end_color='F3490D', fill_type='solid')
                        fail__font = Font(bold=True)
                        wb[each_sheet].cell(row=i[0] + 1, column=5).fill = fail_fill
                        wb[each_sheet].cell(row=i[0] + 1, column=5).font = fail__font
                        Logger("root").getlog().error('data1["describe"]')  # 用例执行失败，写入日志文件中

                else:   # 常规行为操作
                    getattr(keys, data1["action"])(**data1["parameter"])
                Logger("root").getlog().info(data1["describe"])  # 写入日志文件中
    wb.save(file_path)



